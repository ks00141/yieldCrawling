# 구 GW YMS Auto Mail 지원중단으로 O365 스크랩을 위한 자동 로그인 구현

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import chromedriver_autoinstaller
import webbrowser

import datetime
import time
import openpyxl
import os
from openpyxl.chart import BarChart, Reference

try:
    path = chromedriver_autoinstaller.install()

    driver = webdriver.Chrome(path)

except FileNotFoundError as err:
    print("크롬 브라우저를 찾을 수 없습니다. 설치 후 재시도 하시기 바랍니다.")

# o365 URL
URL = 'https://outlook.office.com/'

# chromeDriver LOCAL URI
DRIVRERIURI = 'C:/Users/wisol/Desktop/chromedriver.exe'

# Mail title ROWSTRING
# NS = CSP,WLP / TS = TC-CSP, TC-WLP / PST = PST AOI
TC = '[최종외검][TC-CSP,TC-WLP]최종외검 저수율 리스트'
NS = '[최종외검][CSP,WLP]최종외검 저수율 리스트'
PST = '[PST Map]PST AOI 저수율 리스트'

admin_flag = 0
admin_param = '27'
if admin_flag == 0:
    date = datetime.datetime.now().strftime("[%Y-%m-%d]")
elif admin_flag == 1:
    date = datetime.datetime.now().strftime(f"[%Y-05-{admin_param}]")

dateYM = datetime.datetime.now().strftime("%y%m")



group = [f'{TC} {date}', f'{NS} {date}', f'{PST} {date}']
# webdriver chrome obj 생성
# driver = webdriver.Chrome(DRIVRERIURI)

# o365 URL get request on chrome driver
driver.get(URL)

# page 로딩시간 대기
time.sleep(5)

# ID입력창 가져오기
inputId = driver.find_element_by_id('i0116')

# Daeduck GW ID 입력 / Return 입력
inputId.send_keys('w2200810@wisol.co.kr')
inputId.send_keys(Keys.RETURN)
time.sleep(4)

# ID입력창 가져오기
inputPwd = driver.find_element_by_id('i0118')

# Daeduck GW pwd 입력 / Return 입력
inputPwd.send_keys('K8s,Docker')
inputPwd.send_keys(Keys.RETURN)
time.sleep(2)

# submit btn click
submit = driver.find_element_by_id('idSIButton9')
submit.send_keys(Keys.RETURN)
time.sleep(2)

# 저수율메일함 Web elem click
#cursor = driver.find_element_by_css_selector('#app > div > div.zZJcFiYp1GsQ-Zkcz02eC > div.mXEfuMleN9V2Rx6d6qvsu > div._2aSECY2_aC8BM-pa12gLyl > div > div > div.tQjtZGBXoedSUDzkcRzw5 > div._1mmhFz6xbEHFv6FfTUKPW2 > div > div > div > div:nth-child(4) > div:nth-child(9)')
cursor = driver.find_elements_by_xpath('//*[@id="MainModule"]/div/div/div[1]/div/div/div/div/div[3]/div[12]/div/span[1]')[0];
cursor.click()
time.sleep(5)

def createBarChart(sheet,max_row):
    values = Reference(sheet, min_col=5, min_row=2, max_col=5, max_row=max_row)
    cats = Reference(sheet, min_col=4, min_row=2, max_col=4, max_row=max_row)
    chart = BarChart()
    chart.width = 30
    chart.height = 15
    chart.title = "Yield"
    chart.y_axis.title = 'Yield(%)'
    chart.x_axis.title = 'Wafer ID'
    chart.add_data(values)
    chart.set_categories(cats)
    sheet.add_chart(chart, "G1")

# Mail 공통 클래스 : _1bVZQZoqR8bXQm6sTkfm1W
# Web element list로 받아서
# 하나씩 상대경로로 타고 들어가기(xpath 이용)
# Mail Title = _1bVZQZoqR8bXQm6sTkfm1W > div:nth-child(2) > div > span
# xpath 기본구문
# //tagName[@attribute = 'Value']/tag/tag...
cursor = driver.find_elements_by_xpath('//div[@class = "ZtMcN"]/div/div/div[1]/div[2]/div[2]/div/span')
for i in range(3):
    for title in cursor:
        if(title.text == group[i]):
            print("Matching Succ")
            print(title.text)
            title.click()
            time.sleep(5)
            summary = driver.find_element_by_id('x_summary')
            tr = summary.find_elements_by_xpath('./tbody/tr')[1:]
            if group[i] == f'{TC} {date}':
                print(group[i])
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.cell(row=1, column=1).value = 'NO'
                sheet.cell(row=1, column=2).value = '기종'
                sheet.cell(row=1, column=3).value = 'LOT NO'
                sheet.cell(row=1, column=4).value = 'WAFER ID(MAP)'
                sheet.cell(row=1, column=5).value = '수율(%)'
                max_r = 0
                for i in tr:
                    td = i.find_elements_by_xpath('./td')
                    no = int(td[0].get_attribute('innerHTML'))
                    lotid = td[1].get_attribute('innerHTML')
                    waferid = td[2].get_attribute('innerHTML')
                    device = td[4].get_attribute('innerHTML')
                    yield_ = float(td[6].get_attribute('innerHTML'))
                    sheet.cell(row=no+1,column=1).value = no
                    sheet.cell(row=no+1,column=2).value = device
                    sheet.cell(row=no+1,column=3).value = lotid
                    sheet.cell(row=no+1,column=4).value = waferid
                    sheet.cell(row=no+1,column=4).hyperlink = f'WaferMap://{waferid}:외관검사:최종:DRANK'
                    sheet.cell(row=no+1,column=4).style = 'Hyperlink'
                    sheet.cell(row=no+1,column=5).value = yield_
                    max_r=no+1
                try:
                    createBarChart(sheet=sheet, max_row=max_r)
                    wb.save(f'./저수율list/TC/{dateYM}/TC_저수율list_{date}.xlsx')
                except:
                    print("Create Directory")
                    os.mkdir(f'./저수율list/TC/{dateYM}')
                    wb.save(f'./저수율list/TC/{dateYM}/TC_저수율list_{date}.xlsx')
                print('작업 완료')
                # list_btn = driver.find_element_by_css_selector('#wrap > form > div:nth-child(20) > ul.menuLeft > li:nth-child(8) > a > span')
                # list_btn.click()
                break
            elif group[i] == f'{NS} {date}':
                print(group[i])
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.cell(row=1, column=1).value = 'NO'
                sheet.cell(row=1, column=2).value = '기종'
                sheet.cell(row=1, column=3).value = 'LOT NO'
                sheet.cell(row=1, column=4).value = 'WAFER ID(MAP)'
                sheet.cell(row=1, column=5).value = '수율(%)'
                max_r = 0
                for i in tr:
                    td = i.find_elements_by_xpath('./td')
                    no = int(td[0].get_attribute('innerHTML'))
                    lotid = td[1].get_attribute('innerHTML')
                    waferid = td[2].get_attribute('innerHTML')
                    device = td[4].get_attribute('innerHTML')
                    yield_ = float(td[7].get_attribute('innerHTML'))
                    sheet.cell(row=no+1,column=1).value = no
                    sheet.cell(row=no+1,column=2).value = device
                    sheet.cell(row=no+1,column=3).value = lotid
                    sheet.cell(row=no+1,column=4).value = waferid
                    sheet.cell(row=no+1,column=4).hyperlink = f'WaferMap://{waferid}:외관검사:최종:DRANK'
                    sheet.cell(row=no+1,column=4).style = 'Hyperlink'
                    sheet.cell(row=no+1,column=5).value = yield_
                    max_r = no + 1
                try:
                    createBarChart(sheet=sheet, max_row=max_r)
                    wb.save(f'./저수율list/NS/{dateYM}/NS_저수율list_{date}.xlsx')
                except:
                    os.mkdir(f'./저수율list/NS/{dateYM}')
                    wb.save(f'./저수율list/NS/{dateYM}/NS_저수율list_{date}.xlsx')
                print('작업 완료')
                # list_btn = driver.find_element_by_css_selector('#wrap > form > div:nth-child(20) > ul.menuLeft > li:nth-child(8) > a > span')
                # list_btn.click()
                break
            elif group[i] == f'{PST} {date}':
                print(group[i])
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.cell(row=1, column=1).value = 'NO'
                sheet.cell(row=1, column=2).value = '기종'
                sheet.cell(row=1, column=3).value = 'LOT NO'
                sheet.cell(row=1, column=4).value = 'WAFER ID(MAP)'
                sheet.cell(row=1, column=5).value = '수율(%)'
                max_r = 0
                for i in tr:
                    td = i.find_elements_by_xpath('./td')
                    no = int(td[0].get_attribute('innerHTML'))
                    lotid = td[1].get_attribute('innerHTML')
                    waferid = td[2].get_attribute('innerHTML')
                    device = td[5].get_attribute('innerHTML')
                    yield_ = float(td[10].get_attribute('innerHTML'))
                    sheet.cell(row=no+1,column=1).value = no
                    sheet.cell(row=no+1,column=2).value = device
                    sheet.cell(row=no+1,column=3).value = lotid
                    sheet.cell(row=no+1,column=4).value = waferid
                    sheet.cell(row=no+1,column=4).hyperlink = f'WaferMap://{waferid}:외관검사:PST:FRANK'
                    sheet.cell(row=no+1,column=4).style = 'Hyperlink'
                    sheet.cell(row=no+1,column=5).value = yield_
                    max_r = no + 1
                try:
                    createBarChart(sheet=sheet, max_row=max_r)
                    wb.save(f'./저수율list/PST/{dateYM}/PST_저수율list_{date}.xlsx')
                except:
                    os.mkdir(f'./저수율list/PST/{dateYM}')
                    wb.save(f'./저수율list/PST/{dateYM}/PST_저수율list_{date}.xlsx')
                print('작업 완료')
                # list_btn = driver.find_element_by_css_selector('#wrap > form > div:nth-child(20) > ul.menuLeft > li:nth-child(8) > a > span')
                # list_btn.click()
                break
        else:
            print("Mathcing Fail")
            print(title.text)
            print(group[i])